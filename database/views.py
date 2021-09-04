from django.shortcuts import render, redirect
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import *
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
from .forms import *
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required

def calculate_score(blocks_used):
    weightage = [0, 0, 0, 20, 20, 20, 20, 20]
    best = [0, 5, 5, 5, 5, 5, 5, 5]

    score = 0
    for i in range(1, len(best)):
        if blocks_used[i] != -1:
            score += weightage[i] * best[i] / blocks_used[i]

    return score


@csrf_exempt
def update(request):

    if request.user.is_authenticated:
        payload = json.loads(request.body)

        team = payload['team']
        map_variable = payload['map']
        blocks = payload['blocks']
        code = payload['code']

        instance = Score.objects.get_or_create(team=team)[0]
        instance.user = request.user
        if (instance.blocks[map_variable] > blocks) or (instance.blocks[map_variable] == -1):
            instance.blocks[map_variable] = blocks
            instance.codes[map_variable] = code

            instance.score = calculate_score(instance.blocks)
            instance.time = datetime.now()
            instance.save()

        return JsonResponse({'team': team})
    else:
        return JsonResponse({'team' : None})


@csrf_exempt
def update2(request):

    if request.user.is_authenticated:
        payload = json.loads(request.body)

        team = payload['team']
        level = payload['level']
        code = payload['code']

        instance = Score.objects.get_or_create(team=team)[0]
        instance.user = request.user
        if instance.turtle[level] != 1:
            instance.turtle[level] = 1
            instance.turtle_codes[level] = code
            instance.time2 = datetime.now()

            instance.save()

        return JsonResponse({'team': team})
    else:
        return JsonResponse({'team' : None})




@csrf_exempt
def update3(request):

    if request.user.is_authenticated:
        payload = json.loads(request.body)

        team = payload['team']
        level = payload['level']
        code = payload['code']

        instance = Score.objects.get_or_create(team=team)[0]
        instance.user = request.user
        if instance.movie[level] != 1:
            instance.movie[level] = 1
            instance.movie_codes[level] = code
            instance.time3 = datetime.now()
            instance.save()
        return JsonResponse({'team': team})
    else:
        return JsonResponse({'team' : None})





def create_user_dataset(request):
    if request.method == 'POST':
        form = SheetForm(request.POST, request.FILES)
        if form.is_valid():
            excel_data = request.FILES['file'].read()

            wb = load_workbook(filename=BytesIO(excel_data))
            sheet = wb.active
            row_count = sheet.max_row
            for i in range(row_count):
                roll_no = str(sheet.cell(row=i+2, column=1).value)
                user_email = str(sheet.cell(row=i+2, column=2).value)
                user_password = str(sheet.cell(row=i+2, column=3).value)
                user = User.objects.create_user(username=roll_no, email=user_email, password=user_password)
                user.first_name = str(sheet.cell(row=i+2, column=4).value)
                user.last_name = str(sheet.cell(row=i+2, column=5).value)
                user.roll_no = roll_no
                user.contact = str(sheet.cell(row=i+2, column=6).value)
                user.save()
            return HttpResponse("All users entry have been successfully completed.")
    else:
        form = SheetForm()
    return render(request, 'sheet_upload.html', {'form': form})


def techo_login(request):
    if request.user.is_authenticated:
        return redirect(request.GET.get('next', '/static/index.html'))

    if request.method == "POST":
        technouser = User.objects.filter(roll_no=request.POST['roll_no']).first()
        if technouser is None:
            return render(request, 'login.html', {"messages": [["text-danger", "Roll Number Not Found."]]})
        username = technouser.username
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            print("Used logged in!")
            return redirect(request.GET.get('next', '/static/index.html'))
        else:
            return render(request, 'login.html', {"messages": [["text-danger", "Invalid Credentials."]]})
    return render(request, 'login.html', )


def logout_view(request):
    logout(request)
    return redirect('/')